"""Openpyxl workbook renderer for quotation summary + detail sheets."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

from app.adapters.quotation._xlsx_utils import (
    collect_fieldnames,
    cum_qty_column_index,
    detail_total_column_index,
    excel_sheet_title,
    normalize_row,
    unit_price_column_index,
)
from app.ports.domains.quotation_workbook import QuotationWorkbookRenderPort
from app.ports.dto.quotation_workbook import QuotationWorkbookData, QuotationWorkbookExport

_EXCLUDED_ROW_KEYS: frozenset[str] = frozenset({
    "__root_inv_code",
    "__parent_inv_code",
    "基本用量",
    "供应类型",
    "根父件名称",
    "仓库编码",
    "领料部门",
    "__root_inv_name",
})

_SUMMARY_HEADERS = [
    "\u90e8\u54c1\u7f16\u53f7",
    "\u540d\u79f0",
    "\u6570\u91cf",
    "\u5355\u4ef7",
    "\u91d1\u989d",
    "\u5907\u6ce8",
]


def _sheet_cell_ref(sheet_title: str, col_letter: str, row_num: int) -> str:
    escaped = sheet_title.replace("'", "''")
    return f"='{escaped}'!{col_letter}{row_num}"


class OpenpyxlQuotationWorkbookAdapter(QuotationWorkbookRenderPort):
    def export_workbook(self, workbook_data: QuotationWorkbookData) -> QuotationWorkbookExport:
        from openpyxl import Workbook  # noqa: PLC0415
        from openpyxl.styles import Border, Font, Side  # noqa: PLC0415
        from openpyxl.utils import get_column_letter  # noqa: PLC0415

        wb = Workbook()
        used_titles: set[str] = set()
        bold_font = Font(bold=True)
        thin_side = Side(style="thin")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        summary_ws = wb.active
        summary_ws.title = excel_sheet_title(workbook_data.summary_sheet_name, used_titles)

        detail_sheet_refs: List[Dict[str, Any]] = []
        detail_sheet_refs_by_name: Dict[str, Dict[str, Any]] = {}

        for detail_sheet in workbook_data.detail_sheets:
            ws = wb.create_sheet(title=excel_sheet_title(detail_sheet.sheet_name, used_titles))
            actual_title = ws.title
            if not detail_sheet.rows:
                ws["A1"] = "(no rows)"
                ref = {"title": actual_title}
                detail_sheet_refs.append(ref)
                detail_sheet_refs_by_name[detail_sheet.sheet_name] = ref
                continue
            fieldnames = collect_fieldnames(detail_sheet.rows, _EXCLUDED_ROW_KEYS)
            if not fieldnames:
                ws["A1"] = "(no rows)"
                ref = {"title": actual_title}
                detail_sheet_refs.append(ref)
                detail_sheet_refs_by_name[detail_sheet.sheet_name] = ref
                continue
            ws.append(fieldnames)

            price_col_idx = unit_price_column_index(fieldnames)
            cum_qty_col_idx = cum_qty_column_index(fieldnames)
            total_col_idx = detail_total_column_index(fieldnames)

            data_start_row = 2
            row_num = data_start_row

            for row in detail_sheet.rows:
                normed = normalize_row(row, _EXCLUDED_ROW_KEYS)
                row_values = [normed.get(key) for key in fieldnames]

                if total_col_idx is not None and cum_qty_col_idx is not None and price_col_idx is not None:
                    qty_letter = get_column_letter(cum_qty_col_idx + 1)
                    price_letter = get_column_letter(price_col_idx + 1)
                    row_values[total_col_idx] = f"={qty_letter}{row_num}*{price_letter}{row_num}"

                ws.append(row_values)
                row_num += 1

            ref: Dict[str, Any] = {"title": actual_title}
            if total_col_idx is not None and row_num > data_start_row:
                total_row = [None] * len(fieldnames)
                total_letter = get_column_letter(total_col_idx + 1)
                total_row[total_col_idx] = f"=SUM({total_letter}{data_start_row}:{total_letter}{row_num - 1})"
                ws.append(total_row)
                ref["total_letter"] = total_letter
                ref["total_row_num"] = row_num
            elif price_col_idx is not None and row_num > data_start_row:
                ref["total_letter"] = get_column_letter(price_col_idx + 1)
                ref["total_row_num"] = row_num - 1
            detail_sheet_refs.append(ref)
            detail_sheet_refs_by_name[detail_sheet.sheet_name] = ref

        meta = workbook_data.summary_meta
        summary_ws["C6"] = meta.pricing_title or workbook_data.summary_title
        summary_ws["F3"] = "型号："
        summary_ws["G3"] = meta.model
        summary_ws["F6"] = "报价日期："
        summary_ws["G6"] = meta.quote_date
        summary_ws["E8"] = meta.tax_note
        summary_ws["E9"] = "\u5355\u4f4d\uff1a"
        summary_ws["F9"] = meta.currency_label
        summary_ws["E10"] = "\u603b\u8ba1\uff1a"
        summary_ws["F10"] = meta.grand_total
        summary_ws["B10"] = meta.table_title
        summary_ws.append([])

        summary_header_row = 11
        summary_data_start = summary_header_row + 1
        for col_idx, header in enumerate(_SUMMARY_HEADERS, start=2):
            summary_ws.cell(row=summary_header_row, column=col_idx, value=header)

        row_idx = summary_header_row + 1
        detail_ref_idx = 0
        for row in workbook_data.summary_rows:
            summary_ws.cell(row=row_idx, column=2, value=row.part_no)
            summary_ws.cell(row=row_idx, column=3, value=row.name)
            qty_val: Any = row.quantity_display
            try:
                qty_num = float(qty_val)
                qty_val = int(qty_num) if qty_num == int(qty_num) else qty_num
            except (TypeError, ValueError):
                pass
            summary_ws.cell(row=row_idx, column=4, value=qty_val)

            detail_ref = None
            if row.detail_sheet_name:
                detail_ref = detail_sheet_refs_by_name.get(row.detail_sheet_name)
                if detail_ref is None and detail_ref_idx < len(detail_sheet_refs):
                    detail_ref = detail_sheet_refs[detail_ref_idx]
                detail_ref_idx += 1
            if detail_ref and "total_letter" in detail_ref and "total_row_num" in detail_ref:
                summary_ws.cell(
                    row=row_idx,
                    column=5,
                    value=_sheet_cell_ref(detail_ref["title"], detail_ref["total_letter"], detail_ref["total_row_num"]),
                )
            else:
                summary_ws.cell(row=row_idx, column=5, value=row.unit_price)

            summary_ws.cell(row=row_idx, column=6, value=f"=E{row_idx}*D{row_idx}")
            summary_ws.cell(row=row_idx, column=7, value=row.remark)
            row_idx += 1

        if workbook_data.fixed_charge_rows:
            for row in workbook_data.fixed_charge_rows:
                summary_ws.cell(row=row_idx, column=2, value=row.name)
                summary_ws.cell(row=row_idx, column=3, value=row.quantity_display)
                summary_ws.cell(row=row_idx, column=4, value=row.unit_price)
                summary_ws.cell(row=row_idx, column=5, value=row.amount)
                summary_ws.cell(row=row_idx, column=6, value=row.remark)
                row_idx += 1

        summary_data_end = row_idx - 1
        summary_ws["F10"] = f"=SUM(F{summary_data_start}:F{summary_data_end})" if summary_data_end >= summary_data_start else 0

        for row in summary_ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    cell.font = bold_font

        for cell_range in ("F3:G6", "E8:F10", f"B11:G{row_idx - 1}"):
            for rows in summary_ws[cell_range]:
                for cell in rows:
                    cell.border = thin_border

        bio = BytesIO()
        wb.save(bio)
        return QuotationWorkbookExport(content=bio.getvalue())
