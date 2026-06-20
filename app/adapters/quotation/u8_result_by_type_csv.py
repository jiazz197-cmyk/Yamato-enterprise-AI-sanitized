"""Adapter: ``u8_result_by_type`` → per-type CSV strings and one multi-sheet ``.xlsx``."""

from __future__ import annotations

import csv
import re
from io import BytesIO, StringIO
from typing import Any, Dict, List, Mapping, MutableSet

from app.adapters.quotation._xlsx_utils import (
    collect_fieldnames,
    cum_qty_column_index,
    detail_total_column_index,
    excel_sheet_title,
    normalize_row,
    unit_price_column_index,
)
from app.ports.domains.u8_result_by_type_csv import (
    U8ResultByTypeCsvExport,
    U8ResultByTypeCsvPort,
    U8ResultByTypeXlsxExport,
)

_EXCLUDED_ROW_KEYS: frozenset[str] = frozenset({
    "__root_inv_code",
    "__parent_inv_code",
    "基本用量",
    "供应类型",
    "仓库编码",
    "领料部门",
})
_SAFE_KEY_RE = re.compile(r"[^\w\u4e00-\u9fff\-]+")


def _sanitize_table_key(type_label: str, *, index: int) -> str:
    raw = str(type_label or "").strip() or f"type_{index}"
    cleaned = _SAFE_KEY_RE.sub("_", raw).strip("_")
    return cleaned or f"type_{index}"


def _group_rows_by_sheet(u8_result_by_type: Mapping[str, Any]) -> Dict[str, List[Mapping[str, Any]]]:
    groups = u8_result_by_type.get("items")
    if not isinstance(groups, list):
        return {}

    tables: Dict[str, List[Mapping[str, Any]]] = {}
    used_keys: MutableSet[str] = set()

    for idx, group in enumerate(groups):
        if not isinstance(group, Mapping):
            continue
        type_label = str(group.get("type") or "").strip()
        base_key = _sanitize_table_key(type_label, index=idx)
        key = base_key
        suffix = 1
        while key in used_keys:
            suffix += 1
            key = f"{base_key}_{suffix}"
        used_keys.add(key)

        raw_items = group.get("items")
        rows: List[Mapping[str, Any]] = []
        if isinstance(raw_items, list):
            for item in raw_items:
                if isinstance(item, Mapping):
                    rows.append(item)

        tables[key] = rows

    return tables


def group_u8_bom_rows_by_table_key(
    u8_result_by_type: Mapping[str, Any],
) -> Dict[str, List[Mapping[str, Any]]]:
    """Split ``u8_result_by_type`` into table key → BOM row dicts.

    Keys and insertion order match :meth:`U8ResultByTypeCsvAdapter.export_csv_tables`
    and the worksheet sequence in :meth:`U8ResultByTypeCsvAdapter.export_xlsx_workbook`.
    """

    return _group_rows_by_sheet(u8_result_by_type)


def _rows_to_csv(rows: List[Mapping[str, Any]]) -> str:
    if not rows:
        return ""
    fieldnames = collect_fieldnames(rows, _EXCLUDED_ROW_KEYS)
    if not fieldnames:
        return ""
    buf = StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        if not isinstance(row, Mapping):
            continue
        normed = normalize_row(row, _EXCLUDED_ROW_KEYS)
        writer.writerow({k: normed.get(k) for k in fieldnames})
    return buf.getvalue()


def _rows_by_sheet_to_xlsx(rows_by_sheet: Dict[str, List[Mapping[str, Any]]]) -> bytes:
    from openpyxl import Workbook  # noqa: PLC0415 — optional dependency at runtime
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    wb = Workbook()
    wb.remove(wb.active)
    used_titles: MutableSet[str] = set()

    if not rows_by_sheet:
        ws = wb.create_sheet(title=excel_sheet_title("Empty", used_titles))
        ws["A1"] = "(no grouped types / rows)"
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    for sheet_key, rows in rows_by_sheet.items():
        title = excel_sheet_title(sheet_key, used_titles)
        ws = wb.create_sheet(title=title)
        if not rows:
            ws["A1"] = "(no rows)"
            continue
        fieldnames = collect_fieldnames(rows, _EXCLUDED_ROW_KEYS)
        if not fieldnames:
            ws["A1"] = "(no rows)"
            continue
        ws.append(fieldnames)

        price_col_idx = unit_price_column_index(fieldnames)
        cum_qty_col_idx = cum_qty_column_index(fieldnames)
        total_col_idx = detail_total_column_index(fieldnames)

        data_start_row = 2
        row_num = data_start_row

        for row in rows:
            if not isinstance(row, Mapping):
                continue
            normed = normalize_row(row, _EXCLUDED_ROW_KEYS)
            row_values = [normed.get(k) for k in fieldnames]

            if total_col_idx is not None and cum_qty_col_idx is not None and price_col_idx is not None:
                qty_letter = get_column_letter(cum_qty_col_idx + 1)
                price_letter = get_column_letter(price_col_idx + 1)
                row_values[total_col_idx] = f"={qty_letter}{row_num}*{price_letter}{row_num}"

            ws.append(row_values)
            row_num += 1

        if total_col_idx is not None and row_num > data_start_row:
            total_row = [None] * len(fieldnames)
            total_letter = get_column_letter(total_col_idx + 1)
            total_row[total_col_idx] = f"=SUM({total_letter}{data_start_row}:{total_letter}{row_num - 1})"
            ws.append(total_row)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


class U8ResultByTypeCsvAdapter(U8ResultByTypeCsvPort):
    """Per-type CSV strings and one ``.xlsx`` with a worksheet per type."""

    def export_csv_tables(self, u8_result_by_type: Mapping[str, Any]) -> U8ResultByTypeCsvExport:
        rows_by_sheet = _group_rows_by_sheet(u8_result_by_type)
        tables = {k: _rows_to_csv(v) for k, v in rows_by_sheet.items()}
        return U8ResultByTypeCsvExport(tables=tables)

    def export_xlsx_workbook(self, u8_result_by_type: Mapping[str, Any]) -> U8ResultByTypeXlsxExport:
        try:
            rows_by_sheet = _group_rows_by_sheet(u8_result_by_type)
            return U8ResultByTypeXlsxExport(content=_rows_by_sheet_to_xlsx(rows_by_sheet))
        except ImportError as exc:  # pragma: no cover — exercised when openpyxl missing
            raise ImportError(
                "export_xlsx_workbook requires the 'openpyxl' package (see requirements.txt)."
            ) from exc
