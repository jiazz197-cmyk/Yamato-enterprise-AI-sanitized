from io import BytesIO

from openpyxl import load_workbook

from app.adapters.quotation.workbook import OpenpyxlQuotationWorkbookAdapter
from app.domain.quotation.value_objects import (
    QuotationDetailSheet,
    QuotationSummaryMeta,
    QuotationSummaryRow,
    QuotationWorkbookData,
)


def test_summary_uses_detail_sheet_total_by_sequential_index_and_formula_totals():
    workbook_data = QuotationWorkbookData(
        summary_sheet_name="汇总",
        summary_title="汇总",
        summary_meta=QuotationSummaryMeta(),
        summary_rows=[
            QuotationSummaryRow(
                part_no="A",
                name="外购件",
                quantity_display="1",
                unit_price=12,
                amount=12,
                detail_sheet_name="",
            ),
            QuotationSummaryRow(
                part_no="B",
                name="有明细组件",
                quantity_display="2",
                unit_price=99,
                amount=198,
                detail_sheet_name="组件B",
            ),
        ],
        detail_sheets=[
            QuotationDetailSheet(
                sheet_name="组件B",
                rows=[{"累计用量": 3, "单价": 4, "总价": 12}],
                total_amount=12,
            )
        ],
    )

    export = OpenpyxlQuotationWorkbookAdapter().export_workbook(workbook_data)
    wb = load_workbook(BytesIO(export.content), data_only=False)
    summary_ws = wb["汇总"]

    assert "组件B" in wb.sheetnames
    assert summary_ws["E12"].value == 12
    assert summary_ws["E13"].value == "='组件B'!C3"
    assert summary_ws["F13"].value == "=E13*D13"
    assert summary_ws["F10"].value == "=SUM(F12:F13)"


def test_summary_uses_detail_sheet_price_column_last_row_when_no_total_column():
    workbook_data = QuotationWorkbookData(
        summary_sheet_name="汇总",
        summary_title="汇总",
        summary_meta=QuotationSummaryMeta(),
        summary_rows=[
            QuotationSummaryRow(
                part_no="A",
                name="机架",
                quantity_display="2",
                unit_price=88,
                amount=176,
                detail_sheet_name="机架",
            ),
        ],
        detail_sheets=[
            QuotationDetailSheet(
                sheet_name="机架",
                rows=[
                    {"根父件名称": "机架", "累计用量": 2, "单价": 10, "__root_inv_name": "机架"},
                    {"根父件名称": "机架", "累计用量": "", "单价": 88, "__root_inv_name": "机架"},
                ],
                total_amount=88,
            )
        ],
    )

    export = OpenpyxlQuotationWorkbookAdapter().export_workbook(workbook_data)
    wb = load_workbook(BytesIO(export.content), data_only=False)
    summary_ws = wb["汇总"]

    detail_ws = wb["机架"]

    assert summary_ws["E12"].value == "='机架'!B3"
    assert summary_ws["F12"].value == "=E12*D12"
    assert summary_ws["F10"].value == "=SUM(F12:F12)"
    headers = [cell.value for cell in detail_ws[1]]
    assert "根父件名称" not in headers
    assert "__root_inv_name" not in headers
