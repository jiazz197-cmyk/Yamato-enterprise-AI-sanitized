"""Tests for ``U8ResultByTypeCsvAdapter`` (grouped U8 BOM → CSV + multi-sheet ``.xlsx``).

Artifact export (optional, avoids a second “phase” in the terminal):

- Default when you run this file as a script: **tests only** (one unittest report).
- Add ``--artifacts`` after tests pass: write ``sample.xlsx`` and ``full.xlsx`` under
  ``tests/output/u8_result_by_type/``.
- ``--dump-only``: only write those workbooks, skip tests.

Examples::

    python tests/test_u8_result_by_type_csv.py -v
    python tests/test_u8_result_by_type_csv.py -v --artifacts
    python tests/test_u8_result_by_type_csv.py --dump-only
"""

from __future__ import annotations

import csv
import json
import sys
import unittest
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from app.adapters.quotation.u8_result_by_type_csv import (  # noqa: E402
    U8ResultByTypeCsvAdapter,
    group_u8_bom_rows_by_table_key,
)

_FIXTURE = Path(__file__).resolve().parent / "fixtures" / "u8_result_by_type_sample.json"
_FULL_FIXTURE = Path(__file__).resolve().parent / "fixtures" / "u8_result_by_type_full.json"
_ARTIFACT_DIR = Path(__file__).resolve().parent / "output" / "u8_result_by_type"


def _load_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _write_xlsx_artifacts() -> None:
    """Write ``sample.xlsx`` and ``full.xlsx`` (multi-sheet) into ``tests/output/...``."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("[U8 XLSX] openpyxl is not installed; pip install openpyxl", file=sys.stderr)
        raise SystemExit(2) from None
    adapter = U8ResultByTypeCsvAdapter()
    _ARTIFACT_DIR.mkdir(parents=True, exist_ok=True)
    paths: list[Path] = []
    sample_path = _ARTIFACT_DIR / "sample.xlsx"
    sample_path.write_bytes(adapter.export_xlsx_workbook(_load_json(_FIXTURE)).content)
    paths.append(sample_path)
    if _FULL_FIXTURE.is_file():
        full_path = _ARTIFACT_DIR / "full.xlsx"
        full_path.write_bytes(adapter.export_xlsx_workbook(_load_json(_FULL_FIXTURE)).content)
        paths.append(full_path)
    rels = ", ".join(str(p.relative_to(_PROJECT_ROOT)) for p in paths)
    print(f"\n[U8 XLSX] wrote: {rels}", file=sys.stderr)
    if not _FULL_FIXTURE.is_file():
        print(f"[U8 XLSX] missing {_FULL_FIXTURE.name}; only sample.xlsx", file=sys.stderr)


class TestU8ResultByTypeCsv(unittest.TestCase):
    def test_export_csv_tables_sample_fixture(self):
        payload = _load_json(_FIXTURE)
        adapter = U8ResultByTypeCsvAdapter()
        export = adapter.export_csv_tables(payload)
        rows_by = group_u8_bom_rows_by_table_key(payload)

        self.assertEqual(list(export.tables.keys()), list(rows_by.keys()))
        self.assertEqual(set(export.tables.keys()), set(rows_by.keys()))

        for key, bom_rows in rows_by.items():
            csv_text = export.tables[key]
            if not bom_rows:
                self.assertEqual(csv_text, "", msg=key)
            else:
                self.assertTrue(csv_text.strip(), msg=key)

        # Spot-check rows using stable identifiers from the sample fixture (not type labels).
        rack_key = next(
            (
                k
                for k, rs in rows_by.items()
                if len(rs) == 1 and str(rs[0].get("材料编码（物料编码）")) == "10820900884"
            ),
            None,
        )
        self.assertIsNotNone(rack_key)
        rack_csv = export.tables[rack_key]
        self.assertNotIn("__root_inv_code", rack_csv)
        self.assertNotIn("__parent_inv_code", rack_csv)
        self.assertNotIn("Z0072011", rack_csv)
        self.assertNotIn("u8_parent_inv_codes", rack_csv)

        reader = csv.DictReader(StringIO(rack_csv))
        parsed = list(reader)
        self.assertEqual(len(parsed), 1)
        self.assertEqual(parsed[0]["子件名称"], "振动盘 GB01706G0377c")
        self.assertEqual(parsed[0]["材料编码（物料编码）"], "10820900884")

        funnel_key = next(
            (
                k
                for k, rs in rows_by.items()
                if len(rs) == 2 and str(rs[0].get("材料编码（物料编码）")) == "701100366"
            ),
            None,
        )
        self.assertIsNotNone(funnel_key)
        funnel_rows = list(csv.DictReader(StringIO(export.tables[funnel_key])))
        self.assertEqual(len(funnel_rows), 2)
        self.assertEqual(funnel_rows[0]["供应类型"], "虚拟件")
        self.assertEqual(funnel_rows[1]["子件层级"], "2")

    def test_export_xlsx_workbook_sample_fixture(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed (pip install openpyxl)")
        payload = _load_json(_FIXTURE)
        adapter = U8ResultByTypeCsvAdapter()
        data = adapter.export_xlsx_workbook(payload).content
        self.assertTrue(data.startswith(b"PK"))

        from openpyxl import load_workbook  # noqa: PLC0415 — optional heavy import for this test

        wb = load_workbook(filename=BytesIO(data))
        self.assertEqual(len(wb.sheetnames), len(group_u8_bom_rows_by_table_key(payload)))
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    text = str(cell)
                    self.assertNotIn("__root_inv_code", text)
                    self.assertNotIn("__parent_inv_code", text)

    def test_export_xlsx_empty_items_has_one_sheet(self):
        """Empty ``items`` must still produce a valid workbook (openpyxl requires ≥1 visible sheet)."""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed (pip install openpyxl)")
        adapter = U8ResultByTypeCsvAdapter()
        payload: Dict[str, Any] = {"total": 0, "items": []}
        data = adapter.export_xlsx_workbook(payload).content
        self.assertTrue(data.startswith(b"PK"))

        from openpyxl import load_workbook  # noqa: PLC0415

        wb = load_workbook(filename=BytesIO(data))
        self.assertGreaterEqual(len(wb.sheetnames), 1)

    @unittest.skipUnless(_FULL_FIXTURE.is_file(), f"Missing {_FULL_FIXTURE.name}")
    def test_export_csv_tables_full_user_payload(self):
        """Regression on the full chat ``u8_result_by_type`` payload (``tests/fixtures/u8_result_by_type_full.json``)."""
        payload = _load_json(_FULL_FIXTURE)
        adapter = U8ResultByTypeCsvAdapter()
        export = adapter.export_csv_tables(payload)

        self.assertEqual(payload.get("total"), len(payload.get("items", [])))
        for _key, text in export.tables.items():
            self.assertNotIn("__root_inv_code", text)
            self.assertNotIn("__parent_inv_code", text)
            if not text:
                continue
            for row in csv.DictReader(StringIO(text)):
                self.assertNotIn("__root_inv_code", row)
                self.assertNotIn("__parent_inv_code", row)


if __name__ == "__main__":
    raw = sys.argv[:]
    dump_only = "--dump-only" in raw
    with_artifacts = "--artifacts" in raw
    verbosity = 2 if any(a in ("-v", "--verbose") for a in raw) else 1
    sys.argv = [
        a
        for a in raw
        if a
        not in (
            "--dump-only",
            "--artifacts",
            "-v",
            "--verbose",
        )
    ]
    if dump_only:
        _write_xlsx_artifacts()
        sys.exit(0)

    suite = unittest.TestLoader().loadTestsFromTestCase(TestU8ResultByTypeCsv)
    runner = unittest.TextTestRunner(verbosity=verbosity)
    result = runner.run(suite)
    ok = result.wasSuccessful()
    if ok and with_artifacts:
        _write_xlsx_artifacts()
    sys.exit(0 if ok else 1)
